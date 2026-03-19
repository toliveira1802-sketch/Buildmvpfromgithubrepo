from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

wb = Workbook()
sheet = wb.active
sheet.title = "Colaboradores"

headers = ['id', 'nome', 'cargo', 'setor', 'email', 'username', 'telefone', 'cpf', 'avatar', 'primeiroAcesso', 'nivelAcessoId', 'ativo', 'failedAttempts']
for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

exemplo1 = [1, 'João Silva', 'Mecânico', 'Oficina', 'joao@example.com', 'joao_silva', '11999999999', '12345678901', 'https://avatar.url', True, 2, True, 0]
exemplo2 = [2, 'Maria Santos', 'Consultor', 'Atendimento', 'maria@example.com', 'maria_santos', '11988888888', '98765432101', 'https://avatar.url', True, 1, True, 0]

for col, valor in enumerate(exemplo1, 1):
    sheet.cell(row=2, column=col).value = valor

for col, valor in enumerate(exemplo2, 1):
    sheet.cell(row=3, column=col).value = valor

for col in range(1, len(headers) + 1):
    sheet.column_dimensions[get_column_letter(col)].width = 18

doc_sheet = wb.create_sheet("Documentação")
doc_sheet['A1'] = 'ESTRUTURA DE DADOS - COLABORADORES'
doc_sheet['A1'].font = Font(bold=True, size=14)

docs = [
    ['Campo', 'Tipo', 'Obrigatório', 'Descrição', 'Observações'],
    ['id', 'integer', 'Sim', 'Identificador único (PK)', 'Auto-incremento no banco'],
    ['nome', 'varchar', 'Sim', 'Nome completo do colaborador', 'Sem duplicatas'],
    ['cargo', 'varchar', 'Sim', 'Função do colaborador', 'Valores: Consultor, Mecânico, Gestor, Dev'],
    ['setor', 'varchar', 'Não', 'Setor de atuação', 'Texto livre'],
    ['email', 'varchar', 'Sim', 'E-mail corporativo', 'Único no sistema, não retornar no frontend'],
    ['username', 'varchar', 'Sim', 'Login único', 'Único no sistema'],
    ['telefone', 'varchar', 'Não', 'Telefone de contato', 'Formato: 11999999999'],
    ['cpf', 'varchar', 'Sim', 'CPF do colaborador', 'Nunca retornar em queries públicas'],
    ['avatar', 'varchar', 'Não', 'URL da foto/avatar', 'Link para imagem'],
    ['primeiroAcesso', 'boolean', 'Sim', 'Primeiro acesso?', 'true = deve trocar senha no login'],
    ['nivelAcessoId', 'integer', 'Sim', 'FK para dev_roles', 'Define nível de acesso no sistema'],
    ['ativo', 'boolean', 'Sim', 'Colaborador ativo?', 'true = pode acessar; false = desativa sem deletar'],
    ['failedAttempts', 'integer', 'Sim', 'Tentativas de login falhadas', 'Bloqueio automático após 3 tentativas'],
]

for row, doc_row in enumerate(docs, 1):
    for col, valor in enumerate(doc_row, 1):
        cell = doc_sheet.cell(row=row, column=col)
        cell.value = valor
        if row == 1:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

for col in range(1, 6):
    doc_sheet.column_dimensions[get_column_letter(col)].width = 25

regras_sheet = wb.create_sheet("Regras de Negócio")
regras_sheet['A1'] = 'REGRAS DE NEGÓCIO E VALIDAÇÕES'
regras_sheet['A1'].font = Font(bold=True, size=14)

regras = [
    ['Regra', 'Descrição'],
    ['Unicidade', 'Nome, email e username devem ser únicos - remover registros duplicados'],
    ['Campos Obrigatórios', 'id, nome, cargo, email, username, cpf, primeiroAcesso, nivelAcessoId, ativo, failedAttempts não podem estar vazios'],
    ['Registros Incompletos', 'Remover linhas com campos obrigatórios vazios/null'],
    ['Valores Padrão', 'primeiroAcesso=true, ativo=true, failedAttempts=0 para novos colaboradores'],
    ['Cargo Válido', 'Apenas: Consultor, Mecânico, Gestor, Dev'],
    ['RLS Habilitado', 'Implementar Row Level Security usando nivelAcessoId'],
    ['Segurança', 'Nunca retornar "senha" e "cpf" em queries do frontend'],
    ['Bloqueio de Login', 'Após 3 failedAttempts, usuário fica lockout até admin resetar'],
    ['Troca de Senha', 'Se primeiroAcesso=true, forçar troca no primeiro login'],
]

for row, regra_row in enumerate(regras, 1):
    for col, valor in enumerate(regra_row, 1):
        cell = regras_sheet.cell(row=row, column=col)
        cell.value = valor
        if row == 1:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

for col in range(1, 3):
    regras_sheet.column_dimensions[get_column_letter(col)].width = 40

wb.save('C:\Users\docto\.claude-worktrees\Buildmvpfromgithubrepo\01_colaboradores_estrutura.xlsx')
print("✅ Arquivo criado!")
